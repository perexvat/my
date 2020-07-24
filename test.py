import openpyxl

# читаем excel-файл
wb = openpyxl.load_workbook('d:/example.xlsx')

# печатаем список листов
#sheets = wb.sheetnames
#for sheet in sheets:
 #   print(sheet)

# получаем активный лист
sheet = wb.active

# печатаем значение ячейки A1
#print(sheet['A1'].value)
# печатаем значение ячейки B1
#print(sheet['B1'].value)

#rows = sheet.max_row
#cols = sheet.max_column

#for i in range(1, rows + 1):
#     string = ''
#     for j in range(1, cols + 1):
#         cell = sheet.cell(row = i, column = j)
#         string = string + str(cell.value) + ' '
#     print(string)
# print()
# cell=sheet.cell(row=2, column=40)
# a= str(cell.value)
# #print('в ячейке указано значение')
# #print (a)
# print()

#if (a=='None'):
 #   print ('значение вышло за пределы')
#else:
 #   print(cell.value)
sheet["A"]
for cell in sheet['A']:
    a_zn=(cell.value)
print(a_zn)









